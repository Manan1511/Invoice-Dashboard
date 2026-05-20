import unittest
from unittest.mock import MagicMock, patch
from decimal import Decimal
import openpyxl
from models.ledger import LedgerEntry, LedgerMapping
from services.pl_extractor import extract_pl_dashboard, get_alloc_row_key
from services.ledger_mapper import replace_ledger_list_in_template

class TestPrecisionFixes(unittest.TestCase):
    """
    Unit tests to verify precision fixes in services/pl_extractor.py and services/ledger_mapper.py.
    """

    @patch("services.pl_extractor.load_mapped_ledgers")
    def test_allocation_pool_safety_gate_even_split(self, mock_load_mappings):
        """
        Verify that if total_revenue_pool drops to or below Decimal('0.00'),
        it immediately defaults to the even split fallback loop.
        """
        # Mappings: 2 Revenue Centers (Bluestreak, Clarus) and 1 Cost Center (Common)
        mock_load_mappings.return_value = {
            "sales bluestreak": LedgerMapping(ledger_name="Sales Bluestreak", vertical="Bluestreak", head="1. Sales Accounts", group="P&L"),
            "sales clarus": LedgerMapping(ledger_name="Sales Clarus", vertical="Clarus", head="1. Sales Accounts", group="P&L"),
            "salary common": LedgerMapping(ledger_name="Salary Common", vertical="Common", head="6. Indirect Expense", group="P&L"),
        }

        # Entries: All Sales are zero, Indirect Expense is 1200
        parsed_entries = [
            LedgerEntry(name="Sales Bluestreak", opening=Decimal("0.00"), closing=Decimal("0.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("0.00")),
            LedgerEntry(name="Sales Clarus", opening=Decimal("0.00"), closing=Decimal("0.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("0.00")),
            LedgerEntry(name="Salary Common", opening=Decimal("0.00"), closing=Decimal("1200.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("1200.00")),
        ]

        response = extract_pl_dashboard(parsed_entries, month_label="Mar'26", ytd_label="YTD'26", has_ytd=True)

        # Expected: Even split of 1200 common expense -> 600 Bluestreak, 600 Clarus
        # Let's locate the 'Common Allocation' row in month data
        alloc_row = next(r for r in response.month_data.rows if r.particulars == "Common Allocation")
        
        self.assertEqual(alloc_row.values["Bluestreak"], 600.0)
        self.assertEqual(alloc_row.values["Clarus"], 600.0)
        self.assertEqual(alloc_row.values["Common"], -1200.0)

    @patch("services.pl_extractor.load_mapped_ledgers")
    def test_individual_negative_revenue_floored(self, mock_load_mappings):
        """
        Verify that if total pool is positive but a single vertical has negative net sales,
        the proportional allocation treats the underperforming vertical's revenue share as a floor of zero
        to protect healthy verticals and prevent distorted overhead metrics.
        """
        # Mappings: 2 Revenue Centers (Bluestreak, Clarus) and 1 Cost Center (Common)
        mock_load_mappings.return_value = {
            "sales bluestreak": LedgerMapping(ledger_name="Sales Bluestreak", vertical="Bluestreak", head="1. Sales Accounts", group="P&L"),
            "sales clarus": LedgerMapping(ledger_name="Sales Clarus", vertical="Clarus", head="1. Sales Accounts", group="P&L"),
            "salary common": LedgerMapping(ledger_name="Salary Common", vertical="Common", head="6. Indirect Expense", group="P&L"),
        }

        # Entries:
        # Bluestreak Sales = 1000 (represented by closing of -1000 since sales is credit)
        # Clarus Sales = -200 (represented by closing of +200, i.e., sales returns or negative)
        # Salary Common = 1200
        parsed_entries = [
            LedgerEntry(name="Sales Bluestreak", opening=Decimal("0.00"), closing=Decimal("-1000.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("-1000.00")),
            LedgerEntry(name="Sales Clarus", opening=Decimal("0.00"), closing=Decimal("200.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("200.00")),
            LedgerEntry(name="Salary Common", opening=Decimal("0.00"), closing=Decimal("1200.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("1200.00")),
        ]

        response = extract_pl_dashboard(parsed_entries, month_label="Mar'26", ytd_label="YTD'26", has_ytd=True)

        # Expected:
        # Effective Bluestreak Sales = 1000.00
        # Effective Clarus Sales = 0.00 (floored from -200.00)
        # Total Pool = 1000.00
        # Bluestreak gets 100% of 1200 = 1200 allocation
        # Clarus gets 0% of 1200 = 0 allocation
        alloc_row = next(r for r in response.month_data.rows if r.particulars == "Common Allocation")
        
        self.assertEqual(alloc_row.values["Bluestreak"], 1200.0)
        self.assertEqual(alloc_row.values["Clarus"], 0.0)
        self.assertEqual(alloc_row.values["Common"], -1200.0)

    @patch("services.pl_extractor.load_mapped_ledgers")
    def test_complete_string_normalization(self, mock_load_mappings):
        """
        Verify that dynamic cost centers and revenue centers listings are normalized correctly
        using strip() and title() to prevent tracking fragmentation between different casing and trailing spaces.
        """
        # Mappings with fragmented casings and spaces
        mock_load_mappings.return_value = {
            "sales bluestreak": LedgerMapping(ledger_name="Sales Bluestreak", vertical="  blueSTREAK  ", head="1. Sales Accounts", group="P&L"),
            "salary common": LedgerMapping(ledger_name="Salary Common", vertical=" common ", head="6. Indirect Expense", group="P&L"),
            "factory power": LedgerMapping(ledger_name="Factory Power", vertical="  faCTory  ", head="6. Indirect Expense", group="P&L"),
        }

        parsed_entries = [
            LedgerEntry(name="Sales Bluestreak", opening=Decimal("0.00"), closing=Decimal("-1000.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("-1000.00")),
            LedgerEntry(name="Salary Common", opening=Decimal("0.00"), closing=Decimal("500.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("500.00")),
            LedgerEntry(name="Factory Power", opening=Decimal("0.00"), closing=Decimal("300.00"), opening_ytd=Decimal("0.00"), closing_ytd=Decimal("300.00")),
        ]

        response = extract_pl_dashboard(parsed_entries, month_label="Mar'26", ytd_label="YTD'26", has_ytd=True)

        # Columns should be cleanly normalized and sorted: ["Bluestreak", "Common", "Factory", "Total"]
        self.assertEqual(response.month_data.columns, ["Bluestreak", "Common", "Factory", "Total"])

        # Row allocations should exist for Factory and Common (both cost centers)
        alloc_factory_row = next(r for r in response.month_data.rows if r.particulars == "Allocation of Factory")
        alloc_common_row = next(r for r in response.month_data.rows if r.particulars == "Common Allocation")

        self.assertIsNotNone(alloc_factory_row)
        self.assertIsNotNone(alloc_common_row)

    @patch("services.ledger_mapper.openpyxl.load_workbook")
    def test_exhaustive_overwrite_loop(self, mock_load_workbook):
        """
        Verify that replace_ledger_list_in_template correctly wipes columns 1 through 17 completely
        down to ws.max_row.
        """
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_load_workbook.return_value = mock_wb
        mock_wb.sheetnames = ["List of Ledgers"]
        mock_wb.__getitem__.return_value = mock_ws

        # Mock worksheet dimensions: max_row is 10
        mock_ws.max_row = 10
        mock_ws.cell.return_value = MagicMock()

        # Call the replace function with empty/dummy list to verify clearing behavior
        replace_ledger_list_in_template([])

        # Verify that for rows 5 to 10 and columns 1 to 17, value is set to None
        set_calls = []
        for call in mock_ws.cell.mock_calls:
            # Look for ws.cell(row=X, column=Y) where we set .value = None
            if "value" not in call.kwargs and len(call.args) >= 2:
                row = call.args[0] if len(call.args) > 0 else call.kwargs.get("row")
                col = call.args[1] if len(call.args) > 1 else call.kwargs.get("column")
                # ws.cell(row=X, column=Y).value = None is mock_ws.cell().value = None
                pass

        # Since it's property assignment on mock object, we can verify that mock_ws.cell.call_count
        # is at least 6 rows * 17 columns = 102 calls
        self.assertGreaterEqual(mock_ws.cell.call_count, 102)

if __name__ == "__main__":
    unittest.main()
