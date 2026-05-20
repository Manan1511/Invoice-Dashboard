import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Optional
from models.ledger import LedgerEntry, MappingError

def check_if_tb_has_ytd(parsed_entries: List[LedgerEntry]) -> bool:
    """Checks if the parsed trial balance has YTD values populated."""
    ytd_count = 0
    total_valid = 0
    for entry in parsed_entries:
        if entry.debit is not None or entry.credit is not None or entry.closing is not None:
            total_valid += 1
            if entry.debit_ytd is not None or entry.credit_ytd is not None or entry.closing_ytd is not None:
                ytd_count += 1
                
    if total_valid == 0:
        return False
    # If more than 30% of active entries have YTD data, we assume YTD is supplied by Tally
    return (ytd_count / total_valid) > 0.3

def roll_forward_ytd(
    current_entries: List[LedgerEntry], 
    prior_workbook_path: Optional[str],
    month: int
) -> List[LedgerEntry]:
    """Rolls forward YTD balances by adding current month values to prior month YTD values."""
    from services.ledger_mapper import load_mapped_ledgers, clean_ledger_name
    active_mappings = load_mapped_ledgers()
    if not prior_workbook_path:
        # No prior month workbook provided, initialize YTD values with current month net signed values (e.g. if it's the first month)
        for entry in current_entries:
            if entry.opening_ytd is None:
                entry.opening_ytd = entry.opening_net if entry.opening_net is not None else (entry.opening or Decimal("0.00"))
            if entry.debit_ytd is None:
                entry.debit_ytd = entry.debit_net if entry.debit_net is not None else (entry.debit or Decimal("0.00"))
            if entry.credit_ytd is None:
                entry.credit_ytd = entry.credit_net if entry.credit_net is not None else (entry.credit or Decimal("0.00"))
            if entry.closing_ytd is None:
                entry.closing_ytd = entry.closing_net if entry.closing_net is not None else (entry.closing or Decimal("0.00"))
        return current_entries

    # Load prior month YTD values from its 'List of Ledgers ' sheet
    prior_wb = openpyxl.load_workbook(prior_workbook_path, data_only=True)
    if 'List of Ledgers ' not in prior_wb.sheetnames:
        # Fallback to current values if sheet is missing
        return roll_forward_ytd(current_entries, None, month)
        
    prior_ws = prior_wb['List of Ledgers ']
    prior_ytd_data = {}
    
    # Read columns: Name of Ledger (Col B), Opening YTD (Col N), Debit YTD (Col O), Credit YTD (Col P), Closing YTD (Col Q)
    for r_idx in range(5, prior_ws.max_row + 1):
        name = prior_ws.cell(row=r_idx, column=2).value
        if not name:
            continue
        name_clean = clean_ledger_name(name)
        
        op_ytd = prior_ws.cell(row=r_idx, column=14).value
        deb_ytd = prior_ws.cell(row=r_idx, column=15).value
        cred_ytd = prior_ws.cell(row=r_idx, column=16).value
        clos_ytd = prior_ws.cell(row=r_idx, column=17).value
        
        # Check current active mapping
        mapping = active_mappings.get(name_clean)
        
        if not mapping:
            # Check if this prior ledger has a non-zero balance
            op_ytd_dec = _to_decimal(op_ytd)
            deb_ytd_dec = _to_decimal(deb_ytd)
            cred_ytd_dec = _to_decimal(cred_ytd)
            clos_ytd_dec = _to_decimal(clos_ytd)
            
            if any(val != Decimal("0.00") for val in [op_ytd_dec, deb_ytd_dec, cred_ytd_dec, clos_ytd_dec]):
                raise MappingError(
                    unmapped_ledgers=[name],
                    message=f"CRITICAL ERROR: Prior ledger '{name}' has a non-zero balance "
                            f"but is NOT mapped in your current master template! Please map it first."
                )
            continue
        
        is_pl = False
        if mapping:
            # Check against the standard Heads that strictly belong to the Profit & Loss statement
            pl_heads = {
                "1. Sales Accounts", 
                "2. Indirect Income", 
                "3. Direct Expense", 
                "4. Direct Income", 
                "5. Purchase Accounts", 
                "6. Indirect Expense"
            }
            is_pl = mapping.head in pl_heads
        
        # If it's April (month 4), reset appropriately based on active mapping Group
        if month == 4:
            if is_pl:
                # P&L accounts wipe completely clean
                prior_ytd_data[name_clean] = {
                    "opening_ytd": Decimal("0.00"),
                    "debit_ytd": Decimal("0.00"),
                    "credit_ytd": Decimal("0.00"),
                    "closing_ytd": Decimal("0.00")
                }
            else:
                # Balance Sheet accounts carry forward the CLOSING balance as the new OPENING balance, but reset movements
                clos_ytd_dec = _to_decimal(clos_ytd)
                prior_ytd_data[name_clean] = {
                    "opening_ytd": clos_ytd_dec,
                    "debit_ytd": Decimal("0.00"),
                    "credit_ytd": Decimal("0.00"),
                    "closing_ytd": clos_ytd_dec
                }
        else:
            # Standard mid-year roll forward
            prior_ytd_data[name_clean] = {
                "opening_ytd": _to_decimal(op_ytd),
                "debit_ytd": _to_decimal(deb_ytd),
                "credit_ytd": _to_decimal(cred_ytd),
                "closing_ytd": _to_decimal(clos_ytd)
            }
        
    for entry in current_entries:
        name_clean = clean_ledger_name(entry.name)
        
        # If YTD values are already parsed from TB YTD and valid, keep them!
        if entry.debit_ytd is not None or entry.credit_ytd is not None:
            continue
            
        prior_vals = prior_ytd_data.get(name_clean, {
            "opening_ytd": Decimal("0.00"),
            "debit_ytd": Decimal("0.00"),
            "credit_ytd": Decimal("0.00"),
            "closing_ytd": Decimal("0.00")
        })
        
        # Opening YTD (April 1st) remains the same throughout the fiscal year
        entry.opening_ytd = prior_vals["opening_ytd"]
        
        # YTD Debit = Prior YTD Debit + Current Month Debit (signed net)
        cur_deb = abs(entry.debit_net if entry.debit_net is not None else (entry.debit or Decimal("0.00")))
        entry.debit_ytd = prior_vals["debit_ytd"] + cur_deb
        
        # YTD Credit = Prior YTD Credit + Current Month Credit (signed net)
        cur_cred = abs(entry.credit_net if entry.credit_net is not None else (entry.credit or Decimal("0.00")))
        entry.credit_ytd = prior_vals["credit_ytd"] + cur_cred
        
        # Calculate YTD Closing (signed net)
        entry.closing_ytd = entry.opening_ytd + entry.debit_ytd - entry.credit_ytd
        
    return current_entries

def _to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0.00")
    try:
        if isinstance(val, (int, float, Decimal)):
            return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        s_val = str(val).strip().replace(',', '').replace('₹', '').replace('$', '').replace(' ', '')
        if s_val in ("", "-", "--"):
            return Decimal("0.00")
        return Decimal(s_val).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (ValueError, TypeError, ArithmeticError):
        return Decimal("0.00")

