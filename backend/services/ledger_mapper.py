import openpyxl
import os
import json
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Set, Optional
from services.engine_utils import strict_normalize_ledger_name

@dataclass
class LedgerMapping:
    name: str
    vertical: str
    head: str
    group: str
    classification: str = ""

@dataclass
class CompanyConfiguration:
    mappings: Dict[str, LedgerMapping] = field(default_factory=dict)
    revenue_verticals: Set[str] = field(default_factory=set)
    cost_verticals: Set[str] = field(default_factory=set)


class ValidationError(Exception):
    pass


def parse_mappings_from_excel(uploaded_mapping_path: str) -> CompanyConfiguration:
    """
    Parses mappings from an Excel file without loading them into the global store directly.
    """
    wb = openpyxl.load_workbook(uploaded_mapping_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.strip().lower() in ["list of ledgers", "ledger mapping", "mapping"]), None)
    if not sheet_name:
        sheet_name = wb.active.title
    ws = wb[sheet_name]
    
    header_row_idx = None
    col_map = {}
    for r_idx in range(1, min(ws.max_row + 1, 15)):
        row_vals = [str(ws.cell(row=r_idx, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("ledger" in v for v in row_vals) and any("group" in v for v in row_vals):
            header_row_idx = r_idx
            for col_idx, header_text in enumerate(row_vals, start=1):
                if "ledger" in header_text and "name" in header_text:
                    col_map["ledger_name"] = col_idx
                elif header_text in {"ledger", "particulars", "account"} and "ledger_name" not in col_map:
                    col_map["ledger_name"] = col_idx
                elif "group" in header_text and "ledger_name" not in header_text:
                    col_map["group"] = col_idx
                elif "head" in header_text and "classification" not in header_text:
                    col_map["head"] = col_idx
                elif "classification" in header_text:
                    col_map["classification"] = col_idx
                elif "vertical" in header_text or "business" in header_text:
                    col_map["vertical"] = col_idx
            break
            
    if header_row_idx is None:
        raise ValidationError("Could not locate a valid header row containing Ledger Name and Group.")
    if "ledger_name" not in col_map:
        raise ValidationError("Required column 'Ledger Name' not found.")
        
    config = CompanyConfiguration()
    for r_idx in range(header_row_idx + 1, ws.max_row + 1):
        raw_name = ws.cell(row=r_idx, column=col_map["ledger_name"]).value
        if not raw_name or str(raw_name).strip() == "":
            continue
        clean_name = strict_normalize_ledger_name(raw_name)
        if not clean_name:
            continue
            
        raw_vertical = ws.cell(row=r_idx, column=col_map.get("vertical", -1)).value if "vertical" in col_map else None
        raw_head = ws.cell(row=r_idx, column=col_map.get("head", -1)).value if "head" in col_map else None
        raw_group = ws.cell(row=r_idx, column=col_map.get("group", -1)).value if "group" in col_map else None
        raw_classification = ws.cell(row=r_idx, column=col_map.get("classification", -1)).value if "classification" in col_map else None
        
        vertical_clean = str(raw_vertical).strip().title() if raw_vertical else "Common"
        head_clean = str(raw_head).strip() if raw_head else ""
        group_clean = str(raw_group).strip() if raw_group else ""
        classification_clean = str(raw_classification).strip() if raw_classification else ""
        
        config.mappings[clean_name] = LedgerMapping(
            name=str(raw_name).strip(),
            vertical=vertical_clean,
            head=head_clean,
            group=group_clean,
            classification=classification_clean
        )
    wb.close()
    return config

def load_dynamic_company_mapping(json_path: str) -> CompanyConfiguration:
    if not os.path.exists(json_path):
        return CompanyConfiguration() # Return empty if no mapping exists yet
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    config = CompanyConfiguration()
    has_sales_flags = set()
    has_expenses_flags = set()
    
    for item in data:
        clean_name = strict_normalize_ledger_name(item["name"])
        vertical = str(item.get("vertical", "Common")).strip().title()
        head = str(item.get("head", "")).strip()
        
        config.mappings[clean_name] = LedgerMapping(
            name=str(item["name"]).strip(),
            vertical=vertical,
            head=head,
            group=str(item.get("group", "")).strip(),
            classification=str(item.get("classification", "")).strip()
        )
        
        if "1. Sales Accounts" in head:
            has_sales_flags.add(vertical)
        if any(marker in head for marker in ["Expense", "Cost", "COGS"]):
            has_expenses_flags.add(vertical)
            
    for vert in set(list(has_sales_flags) + list(has_expenses_flags)):
        if vert in has_sales_flags:
            config.revenue_verticals.add(vert)
        else:
            config.cost_verticals.add(vert)
            
    return config

def save_dynamic_company_mapping(json_path: str, new_mappings: List[LedgerMapping]):
    config = load_dynamic_company_mapping(json_path)
    for m in new_mappings:
        clean_name = strict_normalize_ledger_name(m.name)
        config.mappings[clean_name] = m
        
    data = [asdict(m) for m in config.mappings.values()]
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
