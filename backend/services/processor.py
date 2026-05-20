import csv
import io
import os
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List, Optional, Tuple, Union
import openpyxl

# --- CONSTANTS ---
# Precision mapping for currency rounding
CURRENCY_PRECISION = Decimal("0.01")

# Standard Cost Center Identifiers (Case-Insensitive)
STANDARD_COST_CENTERS = {"common", "office", "factory"}

# Tally Accounting Group Heads
HEAD_SALES = "1. Sales Accounts"
HEAD_INDIRECT_INCOME = "2. Indirect Income"
HEAD_DIRECT_EXPENSE = "3. Direct Expense"
HEAD_PURCHASE = "5. Purchase Accounts"
HEAD_INDIRECT_EXPENSE = "6. Indirect Expense"

# Canonical Key Map for Dynamic Mapping Engine
MAPPING_HEADERS = {
    "name": ["ledger name", "name of ledger", "particulars"],
    "vertical": ["business vertical", "vertical"],
    "head": ["head"],
    "group": ["group", "under"]
}

TB_HEADERS = {
    "name": ["particulars", "ledger name", "account"],
    "opening": ["opening bal", "opening balance", "opening"],
    "debit": ["debit bal", "debit balance", "debit"],
    "credit": ["credit bal", "credit balance", "credit"],
    "closing": ["closing bal", "closing balance", "closing"]
}

# Footer rows to ignore during parsing
FOOTER_EXCLUSIONS = {"total", "grand total", "grand", "net profit", "net loss"}


class MappingError(Exception):
    """Raised when an unmapped active ledger is discovered in the Trial Balance."""
    pass


def clean_ledger_name(name: Any) -> str:
    """
    Standardizes ledger names to be immune to invisible whitespace,
    non-breaking spaces, quotation marks, and case sensitivity.
    """
    if name is None:
        return ""
    s = str(name).strip()
    s = unicodedata.normalize("NFKD", s)
    s = s.strip('"').strip("'")
    return s.strip().lower()


def to_decimal(val: Any) -> Decimal:
    """
    Converts any input value safely to a Decimal with standard currency
    truncation (ROUND_HALF_UP to two decimal places).
    """
    if val is None:
        return Decimal("0.00")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val)).quantize(CURRENCY_PRECISION, rounding=ROUND_HALF_UP)
    
    try:
        s_val = str(val).strip().lower()
        if not s_val:
            return Decimal("0.00")
        
        multiplier = Decimal("1")
        if s_val.endswith("cr") or s_val.endswith(" cr"):
            multiplier = Decimal("-1")
            s_val = s_val.rstrip(" cr").strip()
        elif s_val.endswith("dr") or s_val.endswith(" dr"):
            multiplier = Decimal("1")
            s_val = s_val.rstrip(" dr").strip()
            
        if s_val.startswith("(") and s_val.endswith(")"):
            multiplier *= Decimal("-1")
            s_val = s_val[1:-1].strip()
            
        s_val = s_val.replace(",", "").replace("₹", "").replace("$", "").replace(" ", "")
        if s_val in ("", "-", "--"):
            return Decimal("0.00")
            
        return (Decimal(s_val) * multiplier).quantize(CURRENCY_PRECISION, rounding=ROUND_HALF_UP)
    except (ValueError, TypeError, ArithmeticError):
        return Decimal("0.00")


@dataclass
class TrialBalanceRow:
    ledger_name: str
    normalized_name: str
    opening: Decimal
    debit: Decimal
    credit: Decimal
    closing: Decimal
    vertical: str
    head: str
    group: str


class Processor:
    """
    A pure calculation engine for MIS Automation that parses Excel and CSV sources,
    validates ledger mappings, executes P&L mathematical models, and redistributes
    overhead allocations using strict decimal precision.
    """

    def _read_sheet_rows(self, file_source: Union[str, bytes, io.BytesIO]) -> List[List[Any]]:
        """
        Reads rows from a CSV or Excel file source dynamically.
        Handles both file paths and byte streams.
        """
        is_csv = False
        if isinstance(file_source, str):
            is_csv = file_source.lower().endswith(".csv")
            if not os.path.exists(file_source):
                raise FileNotFoundError(f"File not found: {file_source}")
        
        if is_csv:
            with open(file_source, mode="r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                return [row for row in reader]

        # Handle Byte streams / path for Excel
        if isinstance(file_source, (bytes, io.BytesIO)):
            stream = io.BytesIO(file_source) if isinstance(file_source, bytes) else file_source
            wb = openpyxl.load_workbook(stream, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_source, data_only=True)

        try:
            # Find the 'List of Ledgers' sheet or 'TB' sheet if name suggests it, or use the active one
            sheet_names = [s.strip().lower() for s in wb.sheetnames]
            target_idx = 0
            for keyword in ["list of ledgers", "tb", "trial balance"]:
                if keyword in sheet_names:
                    target_idx = sheet_names.index(keyword)
                    break
            
            ws = wb.worksheets[target_idx]
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append(list(r))
            return rows
        finally:
            wb.close()

    def _find_header_row_and_mapping(
        self, rows: List[List[Any]], header_spec: Dict[str, List[str]], required_keys: List[str]
    ) -> Tuple[int, Dict[str, int]]:
        """
        Scans spreadsheet rows dynamically to locate the header row and map column indices.
        Returns a tuple of (header_row_index, canonical_key_to_column_index).
        """
        # Scan up to the first 25 rows for headers
        for r_idx, row in enumerate(rows[:25]):
            col_map = {}
            for c_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().lower()
                for canonical_key, keywords in header_spec.items():
                    if any(kw in cell_str for kw in keywords):
                        col_map[canonical_key] = c_idx

            # Verify if all required keys are successfully mapped
            if all(k in col_map for k in required_keys):
                return r_idx, col_map

        raise ValueError(
            f"Failed to locate a valid header row. Missing some of the required columns: {required_keys}"
        )

    def load_mappings(self, mapping_file: Union[str, bytes, io.BytesIO]) -> Dict[str, dict]:
        """
        Ingests the ledger mapping sheet and registers classifications in memory.
        Returns a dictionary keyed by normalized, cleaned ledger names.
        """
        rows = self._read_sheet_rows(mapping_file)
        if not rows:
            raise ValueError("The mapping file is empty.")

        header_row_idx, col_map = self._find_header_row_and_mapping(
            rows, MAPPING_HEADERS, required_keys=["name", "vertical", "head"]
        )

        mappings = {}
        for row in rows[header_row_idx + 1:]:
            if not row or len(row) <= max(col_map.values()):
                continue

            raw_name = row[col_map["name"]]
            if raw_name is None or str(raw_name).strip() == "":
                continue

            cleaned_name = clean_ledger_name(raw_name)
            vertical = str(row[col_map["vertical"]] or "Common").strip()
            head = str(row[col_map["head"]] or "").strip()
            
            # Group is optional
            group_idx = col_map.get("group")
            group = str(row[group_idx] or "").strip() if group_idx is not None else ""

            mappings[cleaned_name] = {
                "ledger_name": str(raw_name).strip(),
                "vertical": vertical,
                "head": head,
                "group": group
            }

        return mappings

    def parse_tb(
        self, tb_file: Union[str, bytes, io.BytesIO], mappings: Dict[str, dict]
    ) -> List[TrialBalanceRow]:
        """
        Parses Trial Balance data and tags each active ledger with its mapping context.
        Raises MappingError if an active ledger has no mapped vertical/head.
        """
        rows = self._read_sheet_rows(tb_file)
        if not rows:
            raise ValueError("The Trial Balance file is empty.")

        # In TB, 'name' and 'closing' are strictly required. 'opening', 'debit', 'credit' are optional
        header_row_idx, col_map = self._find_header_row_and_mapping(
            rows, TB_HEADERS, required_keys=["name", "closing"]
        )

        parsed_rows = []
        for row in rows[header_row_idx + 1:]:
            if not row or len(row) <= max(col_map.values()):
                continue

            raw_name = row[col_map["name"]]
            if raw_name is None or str(raw_name).strip() == "":
                continue

            cleaned_name = clean_ledger_name(raw_name)
            if cleaned_name in FOOTER_EXCLUSIONS or cleaned_name.startswith(("total", "opening", "closing")):
                continue

            # Load values securely using Decimal
            closing = to_decimal(row[col_map["closing"]])
            
            opening_idx = col_map.get("opening")
            opening = to_decimal(row[opening_idx]) if opening_idx is not None else Decimal("0.00")
            
            debit_idx = col_map.get("debit")
            debit = to_decimal(row[debit_idx]) if debit_idx is not None else Decimal("0.00")
            
            credit_idx = col_map.get("credit")
            credit = to_decimal(row[credit_idx]) if credit_idx is not None else Decimal("0.00")

            # Validate mappings on non-zero balance accounts
            has_balance = any(val != Decimal("0.00") for val in [opening, debit, credit, closing])
            
            if cleaned_name not in mappings:
                if has_balance:
                    raise MappingError(
                        f"CRITICAL ERROR: Ledger '{raw_name}' has a non-zero balance ({closing}) "
                        f"but is NOT mapped in your ledger mappings! Please map this ledger first."
                    )
                # Skip zero-balance unmapped accounts
                continue

            mapping = mappings[cleaned_name]
            parsed_rows.append(
                TrialBalanceRow(
                    ledger_name=mapping["ledger_name"],
                    normalized_name=cleaned_name,
                    opening=opening,
                    debit=debit,
                    credit=credit,
                    closing=closing,
                    vertical=mapping["vertical"],
                    head=mapping["head"],
                    group=mapping["group"]
                )
            )

        return parsed_rows

    def calculate_financials(
        self, parsed_data: List[TrialBalanceRow], stock_data: Dict[str, Dict[str, Decimal]]
    ) -> Dict[str, Dict[str, Any]]:
        """
        Executes granular financial math by grouping Trial Balance ledgers under dynamic verticals.
        Implements high-precision COGS mathematical equations and aggregates P&L metrics.
        """
        # Ensure all stock data keys and vertical names are formatted consistently
        standardized_stock = {k.strip(): v for k, v in stock_data.items()}

        # Discovered Verticals
        verticals = {row.vertical for row in parsed_data if row.vertical}
        if not verticals:
            verticals = {"Common"}

        # Initialize financial structure per vertical
        financials = {}
        for v in verticals:
            financials[v] = {
                "sales": Decimal("0.00"),
                "purchases": Decimal("0.00"),
                "direct_expenses": Decimal("0.00"),
                "opening_stock": Decimal("0.00"),
                "closing_stock": Decimal("0.00"),
                "cogs": Decimal("0.00"),
                "gross_margin": Decimal("0.00"),
                "indirect_income": Decimal("0.00"),
                "indirect_expenses": Decimal("0.00"),
                "pre_allocation_income": Decimal("0.00"),
                "allocated_overheads": {},
                "total_allocated_overhead": Decimal("0.00"),
                "net_income": Decimal("0.00")
            }

        # Step 1: Accumulate transactions based on head definitions
        for row in parsed_data:
            v = row.vertical
            head = row.head

            # Calculate movement (credit is stored as negative inside TB files, so we subtract credit from debit)
            movement = row.closing - row.opening

            # Sign correction for revenue/income heads
            if head in {HEAD_SALES, HEAD_INDIRECT_INCOME}:
                movement = -movement

            if head == HEAD_SALES:
                financials[v]["sales"] += movement
            elif head == HEAD_PURCHASE:
                financials[v]["purchases"] += movement
            elif head == HEAD_DIRECT_EXPENSE:
                financials[v]["direct_expenses"] += movement
            elif head == HEAD_INDIRECT_INCOME:
                financials[v]["indirect_income"] += movement
            elif head == HEAD_INDIRECT_EXPENSE:
                financials[v]["indirect_expenses"] += movement

        # Step 2: Compute Stock & COGS & Gross Margin
        for v, metrics in financials.items():
            # Populate inventory values (fallback to Decimal('0.00'))
            stock_conf = standardized_stock.get(v, {})
            metrics["opening_stock"] = to_decimal(stock_conf.get("opening", Decimal("0.00")))
            metrics["closing_stock"] = to_decimal(stock_conf.get("closing", Decimal("0.00")))

            # COGS = Opening Stock + purchases - Closing Stock
            metrics["cogs"] = metrics["opening_stock"] + metrics["purchases"] - metrics["closing_stock"]

            # Gross Margin = Sales - COGS - Direct Expenses
            metrics["gross_margin"] = metrics["sales"] - metrics["cogs"] - metrics["direct_expenses"]

            # Pre-Allocation Income = Gross Margin + Indirect Income - Indirect Expenses
            metrics["pre_allocation_income"] = (
                metrics["gross_margin"] + metrics["indirect_income"] - metrics["indirect_expenses"]
            )
            # Default final net income to pre-allocation before overhead sweeps
            metrics["net_income"] = metrics["pre_allocation_income"]

        return financials

    def allocate_overheads(self, financial_data: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """
        Performs proportional overhead allocation sweeps from cost centers to operational verticals.
        Ensures a complete cleanup by bringing each Cost Center net income to flat 0.00.
        Falls back to an Even Split among active revenue verticals if total sales are zero.
        """
        all_verticals = set(financial_data.keys())

        # Distinguish between Revenue Verticals and Cost Centers
        cost_centers = {v for v in all_verticals if v.lower() in STANDARD_COST_CENTERS}
        revenue_verticals = all_verticals - cost_centers

        if not revenue_verticals:
            # If no revenue verticals are found, do not apply allocations (cannot distribute overhead)
            return financial_data

        # Sweep each cost center completely
        for cc in sorted(list(cost_centers)):
            metrics = financial_data[cc]
            
            # The overhead expense pool corresponds to the net cost (negative pre-allocation income)
            # A net loss of -5000 means we need to allocate 5000 to operational units
            cc_overhead_pool = -metrics["pre_allocation_income"]
            
            if cc_overhead_pool == Decimal("0.00"):
                continue

            # Calculate total sales across all operational verticals to compute ratio
            total_sales = sum(financial_data[rv]["sales"] for rv in revenue_verticals)

            for rv in sorted(list(revenue_verticals)):
                rv_metrics = financial_data[rv]
                
                # Proportional calculation vs Even Split fallback
                if total_sales > Decimal("0.00"):
                    ratio = rv_metrics["sales"] / total_sales
                    allocated_share = (cc_overhead_pool * ratio).quantize(
                        CURRENCY_PRECISION, rounding=ROUND_HALF_UP
                    )
                else:
                    # Even split fallback
                    allocated_share = (cc_overhead_pool / Decimal(str(len(revenue_verticals)))).quantize(
                        CURRENCY_PRECISION, rounding=ROUND_HALF_UP
                    )

                # Assign allocation records
                rv_metrics["allocated_overheads"][cc] = allocated_share
                rv_metrics["total_allocated_overhead"] += allocated_share
                rv_metrics["net_income"] -= allocated_share

            # Clear Cost Center metrics dynamically (balanced out to exactly zero)
            metrics["allocated_overheads"][cc] = -cc_overhead_pool
            metrics["total_allocated_overhead"] = -cc_overhead_pool
            metrics["net_income"] = Decimal("0.00")

        return financial_data
