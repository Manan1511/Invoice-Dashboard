import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, Any

def construct_dynamic_workbook(extracted_data: Dict[str, Any], output_path: str) -> None:
    """
    Programmatically constructs a spreadsheet cell-by-cell.
    Injects dynamic sum formulas across operational columns.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIS Report"
    
    rev_verts = extracted_data["revenue_verticals"]
    cost_verts = extracted_data["cost_verticals"]
    grid = extracted_data["grid"]
    
    # Check if we need to render Common. If no rev or cost, we must render Common.
    # Otherwise render operational + Common just in case, but typically we only need them if they have balances.
    # For now, let's render Revenue + Cost + Common
    render_cols = rev_verts + cost_verts + ['Common']
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 1. Header Row
    ws.cell(row=1, column=1, value="Particulars").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    
    col_idx = 2
    for vert in render_cols:
        cell = ws.cell(row=1, column=col_idx, value=vert)
        cell.font = header_font
        cell.fill = header_fill
        col_idx += 1
        
    total_col_idx = col_idx
    total_cell = ws.cell(row=1, column=total_col_idx, value="Total")
    total_cell.font = header_font
    total_cell.fill = header_fill
    
    # Adjust column width
    ws.column_dimensions['A'].width = 40
    for c in range(2, total_col_idx + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
        
    current_row = 2
    
    def write_row(title: str, data_dict: dict, is_bold: bool = False):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title)
        if is_bold:
            ws.cell(row=current_row, column=1).font = bold_font
            
        c_idx = 2
        for vert in render_cols:
            val = float(data_dict.get(vert, 0.00))
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.number_format = '#,##0.00'
            if is_bold:
                cell.font = bold_font
            c_idx += 1
            
        # Write Excel SUM formula
        start_col_letter = get_column_letter(2)
        end_col_letter = get_column_letter(total_col_idx - 1)
        sum_formula = f"=SUM({start_col_letter}{current_row}:{end_col_letter}{current_row})"
        
        sum_cell = ws.cell(row=current_row, column=total_col_idx, value=sum_formula)
        sum_cell.number_format = '#,##0.00'
        if is_bold:
            sum_cell.font = bold_font
            
        current_row += 1

    # Render groups
    write_row('1. Sales Accounts', grid['1. Sales Accounts'], is_bold=True)
    write_row('2. Less: COGS', grid['2. Less: COGS'])
    write_row('3. Direct Expense', grid['3. Direct Expense'])
    write_row('4. Gross Margin', grid['4. Gross Margin'], is_bold=True)
    
    # Spacing
    current_row += 1
    
    write_row('5. Indirect Income', grid['5. Indirect Income'])
    write_row('6. Net Allocable Income', grid['6. Net Allocable Income'], is_bold=True)
    
    current_row += 1
    
    write_row('7. Indirect Expense', grid['7. Indirect Expense'])
    
    # Allocations
    if 'Allocations' in grid and grid['Allocations']:
        current_row += 1
        ws.cell(row=current_row, column=1, value="Allocations").font = bold_font
        current_row += 1
        for alloc_name, alloc_data in grid['Allocations'].items():
            write_row(alloc_name, alloc_data)
            
    current_row += 1
    write_row('Total Indirect Costs', grid['Total Indirect Costs'], is_bold=True)
    
    current_row += 1
    write_row('Net Profit', grid['Net Profit'], is_bold=True)
    
    wb.save(output_path)
    wb.close()
