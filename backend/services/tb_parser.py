import openpyxl
import unicodedata
from typing import List, Dict, Optional, Tuple
from decimal import Decimal, ROUND_HALF_UP
from models.ledger import LedgerEntry
from services.ledger_mapper import strict_normalize_ledger_name

CURRENCY_PRECISION = Decimal("0.01")

def parse_tally_tb(file_path: str) -> List[LedgerEntry]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 1. Parse Monthly TB (from 'TB ' or first sheet)
    monthly_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ['tb', 'trial balance']:
            monthly_sheet = wb[name]
            break
    if not monthly_sheet:
        monthly_sheet = wb.active
        
    monthly_entries = _parse_sheet_rows(monthly_sheet, is_ytd=False)
    
    # 2. Parse YTD TB if exists
    ytd_sheet = None
    for name in wb.sheetnames:
        if 'ytd' in name.strip().lower():
            ytd_sheet = wb[name]
            break
            
    if ytd_sheet:
        ytd_entries = _parse_sheet_rows(ytd_sheet, is_ytd=True)
        # Merge YTD into monthly entries
        for name, ytd_entry in ytd_entries.items():
            if name in monthly_entries:
                monthly_entries[name].opening_ytd = ytd_entry.opening_ytd
                monthly_entries[name].debit_ytd = ytd_entry.debit_ytd
                monthly_entries[name].credit_ytd = ytd_entry.credit_ytd
                monthly_entries[name].closing_ytd = ytd_entry.closing_ytd
            else:
                # Ledger only in YTD sheet (rare, but possible)
                monthly_entries[name] = ytd_entry
                
    return list(monthly_entries.values())

def _parse_sheet_rows(sheet, is_ytd: bool) -> Dict[str, LedgerEntry]:
    entries = {}
    
    # Find the header row and particulars column dynamically
    header_row_idx = -1
    particulars_col = 1
    for r_idx in range(1, 21):
        for c_idx in range(1, min(sheet.max_column + 1, 15)):
            val = sheet.cell(row=r_idx, column=c_idx).value
            if val and str(val).strip().lower() in ['particulars', 'ledger name', 'particular']:
                header_row_idx = r_idx
                particulars_col = c_idx
                break
        if header_row_idx != -1:
            break
            
    if header_row_idx == -1:
        # Fallback to row 4 as standard in Tally template
        header_row_idx = 4
        particulars_col = 1
        
    # Read columns from the header row
    cols = []
    max_cols_to_scan = max(sheet.max_column, 15)
    for c_idx in range(1, max_cols_to_scan + 1):
        val = sheet.cell(row=header_row_idx, column=c_idx).value
        cols.append(str(val).strip() if val else "")
        
    # Establish safe defaults relative to the dynamic particulars_col
    opening_col = particulars_col + 1
    debit_col = particulars_col + 2
    credit_col = particulars_col + 3
    closing_col = particulars_col + 4
    
    opening_net_col = particulars_col + 6
    debit_net_col = particulars_col + 7
    credit_net_col = particulars_col + 8
    closing_net_col = particulars_col + 9

    # Scan column headers to override standard and net columns dynamically
    for c_idx in range(1, len(cols) + 1):
        col_name = cols[c_idx - 1]
        if not col_name:
            continue
        col_lower = col_name.lower()
        
        # If YTD, check if YTD values are in cols (Opening YTD, Debit YTD, etc.)
        if is_ytd:
            if 'opening ytd' in col_lower or ('opening' in col_lower and c_idx >= particulars_col + 6):
                opening_col = c_idx
            elif 'debit ytd' in col_lower or ('debit' in col_lower and c_idx >= particulars_col + 6):
                debit_col = c_idx
            elif 'credit ytd' in col_lower or ('credit' in col_lower and c_idx >= particulars_col + 6):
                credit_col = c_idx
            elif 'closing ytd' in col_lower or ('closing' in col_lower and c_idx >= particulars_col + 6):
                closing_col = c_idx
        else:
            if c_idx >= particulars_col + 6:
                if col_lower == 'opening':
                    opening_net_col = c_idx
                elif col_lower == 'debit':
                    debit_net_col = c_idx
                elif col_lower == 'credit':
                    credit_net_col = c_idx
                elif col_lower == 'closing':
                    closing_net_col = c_idx

    # Use a set for O(1) lookup and exact matching on cleaned names
    FOOTER_EXCLUSIONS = {'total', 'grand total', 'grand', 'net profit', 'net loss'}

    # Parse rows below the header
    consecutive_empty = 0
    for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
        raw_name = sheet.cell(row=r_idx, column=particulars_col).value
        if not raw_name or str(raw_name).strip() == "":
            consecutive_empty += 1
            if consecutive_empty > 50:
                break
            continue
        consecutive_empty = 0
            
        clean_name = strict_normalize_ledger_name(raw_name)
        if not clean_name:
            continue
        
        # Check for exact matches to avoid dropping legitimate ledgers
        if clean_name in FOOTER_EXCLUSIONS:
            continue
        
        # Read cell values
        op_val = _clean_decimal(sheet.cell(row=r_idx, column=opening_col).value)
        deb_val = _clean_decimal(sheet.cell(row=r_idx, column=debit_col).value)
        cred_val = _clean_decimal(sheet.cell(row=r_idx, column=credit_col).value)
        clos_val = _clean_decimal(sheet.cell(row=r_idx, column=closing_col).value)
        
        # Standardize display name casing but keep formatting
        display_name = unicodedata.normalize("NFKD", str(raw_name)).strip().strip('"').strip("'").strip()

        if is_ytd:
            # Skip if all values are None/Zero
            if not any([op_val, deb_val, cred_val, clos_val]):
                continue
            entries[clean_name] = LedgerEntry(
                name=display_name,
                opening_ytd=op_val,
                debit_ytd=deb_val,
                credit_ytd=cred_val,
                closing_ytd=clos_val
            )
        else:
            op_net = _clean_decimal(sheet.cell(row=r_idx, column=opening_net_col).value)
            deb_net = _clean_decimal(sheet.cell(row=r_idx, column=debit_net_col).value)
            cred_net = _clean_decimal(sheet.cell(row=r_idx, column=credit_net_col).value)
            clos_net = _clean_decimal(sheet.cell(row=r_idx, column=closing_net_col).value)
            
            # Skip if all values are None/Zero
            if not any([op_val, deb_val, cred_val, clos_val, op_net, deb_net, cred_net, clos_net]):
                continue
                
            entries[clean_name] = LedgerEntry(
                name=display_name,
                opening=op_val,
                debit=deb_val,
                credit=cred_val,
                closing=clos_val,
                opening_net=op_net,
                debit_net=deb_net,
                credit_net=cred_net,
                closing_net=clos_net
            )
            
    return entries

def _clean_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float, Decimal)):
            return Decimal(str(val)).quantize(CURRENCY_PRECISION, rounding=ROUND_HALF_UP)
            
        s_val = str(val).strip().lower()
        if not s_val:
            return None
            
        # Determine sign based on Tally suffix
        multiplier = Decimal("1")
        if s_val.endswith('cr') or s_val.endswith(' cr'):
            multiplier = Decimal("-1")
            if s_val.endswith(' cr'):
                s_val = s_val[:-3].strip()
            elif s_val.endswith('cr'):
                s_val = s_val[:-2].strip()
        elif s_val.endswith('dr') or s_val.endswith(' dr'):
            multiplier = Decimal("1")
            if s_val.endswith(' dr'):
                s_val = s_val[:-3].strip()
            elif s_val.endswith('dr'):
                s_val = s_val[:-2].strip()

        # Handle parentheses for negative numbers (e.g., (1,234.00))
        if s_val.startswith('(') and s_val.endswith(')'):
            multiplier *= Decimal("-1")
            s_val = s_val[1:-1].strip()

        # Strip commas, currency symbols, and spaces
        s_val = s_val.replace(',', '').replace('₹', '').replace('$', '').replace(' ', '')
        
        if s_val == "" or s_val == "-" or s_val == "--":
            return None
            
        return (Decimal(s_val) * multiplier).quantize(CURRENCY_PRECISION, rounding=ROUND_HALF_UP)
    except (ValueError, TypeError, ArithmeticError):
        return None
