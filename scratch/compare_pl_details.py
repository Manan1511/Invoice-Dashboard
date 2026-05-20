import openpyxl
import os

man_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_April 2025.xlsx"
gen_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/scratch/output_April_2025.xlsx"

wb_man = openpyxl.load_workbook(man_path, data_only=True)
wb_gen = openpyxl.load_workbook(gen_path, data_only=True)

sheet_man = wb_man['P&L']
sheet_gen = wb_gen['P&L']

def map_columns(sheet) -> dict:
    monthly_cols = {}
    is_ytd = False
    for col in range(2, 30):
        r5_val = sheet.cell(row=5, column=col).value
        if r5_val and ('ytd' in str(r5_val).lower() or 'april to' in str(r5_val).lower()):
            is_ytd = True
        r6_val = sheet.cell(row=6, column=col).value
        if r6_val:
            name = str(r6_val).strip()
            if not is_ytd:
                monthly_cols[name] = col
    return monthly_cols

man_cols = map_columns(sheet_man)
gen_cols = map_columns(sheet_gen)

print("Manual Monthly Verticals:", man_cols)
print("Generated Monthly Verticals:", gen_cols)

# We want to compare Bluestreak and Clarus row-by-row
for vert in ['Bluestreak', 'Clarus']:
    print(f"\n==========================================")
    print(f"VERTICAL: {vert}")
    print(f"==========================================")
    print(f"{'Particulars':<40} | {'Manual':<15} | {'Generated':<15} | {'Diff':<15}")
    print("-" * 92)
    
    col_man = man_cols[vert]
    col_gen = gen_cols[vert]
    
    for r_gen in range(8, 100):
        lbl_gen = sheet_gen.cell(row=r_gen, column=1).value
        if not lbl_gen or str(lbl_gen).strip() == "":
            continue
            
        # Search for corresponding row in manual
        r_man = None
        for r_search in range(8, 110):
            lbl_man = sheet_man.cell(row=r_search, column=1).value
            if lbl_man and str(lbl_man).strip().lower() == str(lbl_gen).strip().lower():
                r_man = r_search
                break
                
        if r_man is None:
            # Not found in manual P&L
            val_gen = sheet_gen.cell(row=r_gen, column=col_gen).value
            print(f"{str(lbl_gen)[:40]:<40} | {'[NOT FOUND]':<15} | {str(val_gen):<15} | -")
            continue
            
        val_man = sheet_man.cell(row=r_man, column=col_man).value
        val_gen = sheet_gen.cell(row=r_gen, column=col_gen).value
        
        # Calculate diff if numbers
        diff = ""
        v_man = 0.0
        v_gen = 0.0
        if isinstance(val_man, (int, float)):
            v_man = float(val_man)
        if isinstance(val_gen, (int, float)):
            v_gen = float(val_gen)
            
        if abs(v_man - v_gen) > 0.01:
            diff = f"{abs(v_man - v_gen):.2f}"
            
        print(f"{str(lbl_gen)[:40]:<40} | {str(val_man):<15} | {str(val_gen):<15} | {diff:<15}")
