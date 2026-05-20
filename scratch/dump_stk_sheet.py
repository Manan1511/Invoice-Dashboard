import openpyxl
import os

path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_April 2025.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['Stk ']
print(f"Sheet Name: 'Stk '")
print(f"Max rows: {sheet.max_row}")

for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    if any(x is not None for x in row_vals):
        print(f"Row {r}: {row_vals}")
