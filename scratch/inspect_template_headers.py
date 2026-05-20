import openpyxl
import os

path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/templates/MIS_template.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['P&L']

print("Template Row 5 Headers:")
print([sheet.cell(row=5, column=c).value for c in range(1, 30)])
print("\nTemplate Row 6 Headers:")
print([sheet.cell(row=6, column=c).value for c in range(1, 30)])
