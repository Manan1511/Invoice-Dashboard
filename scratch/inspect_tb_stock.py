import openpyxl
import os

files = [
    "1. MIS_April 2025.xlsx",
    "MIS_May 2025.xlsx",
    "1. MIS_June 2025.xlsx"
]

for filename in files:
    path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"=== {filename} ===")
    if 'TB ' in wb.sheetnames:
        sheet = wb['TB ']
        for r in range(1, sheet.max_row + 1):
            val = sheet.cell(row=r, column=1).value
            if val and 'stock' in str(val).lower():
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 11)]
                print(f"Row {r}: {row_vals}")
    print()
