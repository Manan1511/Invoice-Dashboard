import openpyxl
import os

path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_April 2025.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['P&L']

print("Row 4 Headers:")
print([sheet.cell(row=4, column=c).value for c in range(1, 30)])
print("\nRow 5 Headers:")
print([sheet.cell(row=5, column=c).value for c in range(1, 30)])
print("\nRow 6 Headers:")
print([sheet.cell(row=6, column=c).value for c in range(1, 30)])
