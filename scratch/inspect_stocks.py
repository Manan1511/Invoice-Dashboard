import openpyxl
import os

files = [
    "1. MIS_April 2025.xlsx",
    "MIS_May 2025.xlsx",
    "1. MIS_June 2025.xlsx"
]

for filename in files:
    path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"=== {filename} ===")
    
    # Let's inspect 'Stk ' sheet
    if 'Stk ' in wb.sheetnames:
        sheet = wb['Stk ']
        print("Stk sheet rows:")
        for r in range(1, 20):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 8)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")
                
    # Let's also check if 'COGS' sheet has anything interesting
    if 'COGS' in wb.sheetnames:
        sheet = wb['COGS']
        print("COGS sheet rows:")
        for r in range(1, 15):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 5)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")
    print()
