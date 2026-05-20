import openpyxl

template_path = "templates/MIS_template.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb['P&L']

print("=== Template P&L Row Labels ===")
for r in range(1, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val:
        print(f"Row {r:<3} | {val}")
