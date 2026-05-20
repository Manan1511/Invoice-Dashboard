import openpyxl

man_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_April 2025.xlsx"
wb = openpyxl.load_workbook(man_path, data_only=True)
sheet = wb['P&L']

print("=== April 2025 Manual P&L (Rows 15 to 35) ===")
for r in range(15, 36):
    lbl = sheet.cell(row=r, column=1).value
    val_b = sheet.cell(row=r, column=2).value # Bluestreak
    val_c = sheet.cell(row=r, column=3).value # Clarus
    val_h = sheet.cell(row=r, column=8).value # Common
    val_i = sheet.cell(row=r, column=9).value # Total without share trading
    print(f"Row {r:<2} | {str(lbl or ''):<35} | Bluestreak={val_b} | Clarus={val_c} | Common={val_h} | Total={val_i}")
