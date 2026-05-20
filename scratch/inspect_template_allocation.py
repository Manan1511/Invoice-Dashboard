import openpyxl

template_path = "templates/MIS_template.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=False)

if 'P&L' in wb.sheetnames:
    sheet = wb['P&L']
    print("=== Template Allocation Row Formulas (P&L Sheet) ===")
    for r in range(70, 81):
        lbl = sheet.cell(row=r, column=1).value
        print(f"\nRow {r} | Label: {lbl}")
        for c in range(1, 14):
            h_lbl = sheet.cell(row=6, column=c).value
            form = sheet.cell(row=r, column=c).value
            if form is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"  Col {col_letter} ({h_lbl}): Formula={form}")
