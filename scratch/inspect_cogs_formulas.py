import openpyxl
import os

files = [
    "1. MIS_April 2025.xlsx",
    "MIS_May 2025.xlsx",
    "1. MIS_June 2025.xlsx"
]

for filename in files:
    path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", filename)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    
    if 'COGS' in wb_f.sheetnames:
        sheet_f = wb_f['COGS']
        sheet_v = wb_v['COGS']
        print(f"=== {filename} - COGS ===")
        for r in range(1, 30):
            row_lbl = sheet_v.cell(row=r, column=1).value
            if any(sheet_v.cell(row=r, column=c).value is not None for c in range(1, 15)):
                row_vals = []
                for c in range(1, 15):
                    v = sheet_v.cell(row=r, column=c).value
                    f = sheet_f.cell(row=r, column=c).value
                    if v is not None or f is not None:
                        col_letter = openpyxl.utils.get_column_letter(c)
                        row_vals.append(f"{col_letter}: [Val={v}, Formula={f}]")
                print(f"Row {r}: " + ", ".join(row_vals))
    print()
