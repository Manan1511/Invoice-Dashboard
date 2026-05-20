import openpyxl

man_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_April 2025.xlsx"
wb_f = openpyxl.load_workbook(man_path, data_only=False)
wb_v = openpyxl.load_workbook(man_path, data_only=True)

sheet_f = wb_f['P&L']
sheet_v = wb_v['P&L']

print("=== April 2025 Manual P&L - Rows 15 to 19 Formula Inspect ===")
for r in range(15, 20):
    lbl = sheet_v.cell(row=r, column=1).value
    print(f"\nRow {r} | Label: {lbl}")
    for c in range(1, 12):
        h_lbl = sheet_v.cell(row=6, column=c).value
        val = sheet_v.cell(row=r, column=c).value
        form = sheet_f.cell(row=r, column=c).value
        if val is not None or form is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  Col {col_letter} ({h_lbl}): Val={val}, Formula={form}")
