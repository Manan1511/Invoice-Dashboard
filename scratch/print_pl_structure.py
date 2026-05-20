import openpyxl
import os

gen_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/scratch/output_June_2025.xlsx"
man_path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_June 2025.xlsx"

wb_gen = openpyxl.load_workbook(gen_path, read_only=True)
wb_man = openpyxl.load_workbook(man_path, read_only=True)

sheet_gen = wb_gen['P&L']
sheet_man = wb_man['P&L']

print(f"{'Row':<5} | {'Generated Label':<40} | {'Manual Label':<40}")
print("-" * 90)
for r in range(70, min(sheet_gen.max_row, sheet_man.max_row) + 1):
    lbl_gen = sheet_gen.cell(row=r, column=1).value
    lbl_man = sheet_man.cell(row=r, column=1).value
    if lbl_gen or lbl_man:
        print(f"{r:<5} | {str(lbl_gen or ''):<40} | {str(lbl_man or ''):<40}")
