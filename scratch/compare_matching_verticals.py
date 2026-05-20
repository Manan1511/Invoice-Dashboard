import os
import openpyxl

months_config = [
    {
        "name": "April 2025",
        "manual_file": "1. MIS_April 2025.xlsx",
        "generated_file": "scratch/output_April_2025.xlsx",
    },
    {
        "name": "May 2025",
        "manual_file": "MIS_May 2025.xlsx",
        "generated_file": "scratch/output_May_2025.xlsx",
    },
    {
        "name": "June 2025",
        "manual_file": "1. MIS_June 2025.xlsx",
        "generated_file": "scratch/output_June_2025.xlsx",
    }
]

def map_columns(sheet) -> dict:
    """Scans row 6 and row 5 to map vertical names to columns for Monthly and YTD."""
    # Row 5 contains monthly vs YTD headers
    # Row 6 contains vertical headers
    monthly_cols = {}
    ytd_cols = {}
    
    # We scan columns 2 to 30
    is_ytd_section = False
    
    for col in range(2, 30):
        # Check row 5 to see if YTD section has started
        r5_val = sheet.cell(row=5, column=col).value
        if r5_val and ('ytd' in str(r5_val).lower() or 'april to' in str(r5_val).lower()):
            is_ytd_section = True
            
        r6_val = sheet.cell(row=6, column=col).value
        if r6_val:
            name = str(r6_val).strip()
            if not is_ytd_section:
                monthly_cols[name] = col
            else:
                ytd_cols[name] = col
                
    return {"monthly": monthly_cols, "ytd": ytd_cols}

def compare_verticals():
    print("==================================================")
    print("DYNAMIC VERTICAL MAPPING COMPARISON")
    print("==================================================")
    
    for cfg in months_config:
        man_path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", cfg["manual_file"])
        gen_path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", cfg["generated_file"])
        
        print(f"\n--- Analyzing {cfg['name']} ---")
        
        wb_gen = openpyxl.load_workbook(gen_path, data_only=True)
        wb_man = openpyxl.load_workbook(man_path, data_only=True)
        
        sheet_gen = wb_gen['P&L']
        sheet_man = wb_man['P&L']
        
        # Map columns
        gen_maps = map_columns(sheet_gen)
        man_maps = map_columns(sheet_man)
        
        print("Generated Monthly Columns Mapped:", gen_maps["monthly"])
        print("Manual Monthly Columns Mapped:   ", man_maps["monthly"])
        print("Generated YTD Columns Mapped:    ", gen_maps["ytd"])
        print("Manual YTD Columns Mapped:       ", man_maps["ytd"])
        
        # We will compare shared verticals
        shared_monthly = set(gen_maps["monthly"].keys()).intersection(man_maps["monthly"].keys())
        shared_ytd = set(gen_maps["ytd"].keys()).intersection(man_maps["ytd"].keys())
        
        print(f"Shared Monthly Verticals to compare: {list(shared_monthly)}")
        print(f"Shared YTD Verticals to compare:     {list(shared_ytd)}")
        
        mismatches = 0
        total_checks = 0
        
        # Compare row 8 to row 100
        for r in range(8, 100):
            particulars_gen = sheet_gen.cell(row=r, column=1).value
            particulars_man = sheet_man.cell(row=r, column=1).value
            
            # Skip empty rows or title rows
            if not particulars_gen or str(particulars_gen).strip() == "":
                continue
            if particulars_gen != particulars_man:
                # Row alignment issue (e.g. if one template has more rows)
                # Let's search for the row with same particulars in manual sheet to align them!
                aligned_r_man = None
                for search_r in range(8, 110):
                    if sheet_man.cell(row=search_r, column=1).value == particulars_gen:
                        aligned_r_man = search_r
                        break
                if aligned_r_man is None:
                    # Particulars row does not exist in manual sheet
                    continue
            else:
                aligned_r_man = r
                
            # Compare monthly columns
            for vert in shared_monthly:
                col_gen = gen_maps["monthly"][vert]
                col_man = man_maps["monthly"][vert]
                
                val_gen = sheet_gen.cell(row=r, column=col_gen).value
                val_man = sheet_man.cell(row=aligned_r_man, column=col_man).value
                
                total_checks += 1
                if isinstance(val_gen, (int, float)) and isinstance(val_man, (int, float)):
                    if abs(val_gen - val_man) > 0.5:
                        mismatches += 1
                        print(f"  Monthly Mismatch at Row {r} ({particulars_gen}) for {vert}: Generated={val_gen:.2f}, Manual={val_man:.2f} (Diff={abs(val_gen-val_man):.2f})")
                elif val_gen != val_man:
                    if (val_gen is None and val_man == 0) or (val_gen == 0 and val_man is None):
                        continue
                    mismatches += 1
                    print(f"  Monthly Text Mismatch at Row {r} ({particulars_gen}) for {vert}: Generated={repr(val_gen)}, Manual={repr(val_man)}")
                    
            # Compare YTD columns
            for vert in shared_ytd:
                col_gen = gen_maps["ytd"][vert]
                col_man = man_maps["ytd"][vert]
                
                val_gen = sheet_gen.cell(row=r, column=col_gen).value
                val_man = sheet_man.cell(row=aligned_r_man, column=col_man).value
                
                total_checks += 1
                if isinstance(val_gen, (int, float)) and isinstance(val_man, (int, float)):
                    if abs(val_gen - val_man) > 0.5:
                        mismatches += 1
                        print(f"  YTD Mismatch at Row {r} ({particulars_gen}) for {vert}: Generated={val_gen:.2f}, Manual={val_man:.2f} (Diff={abs(val_gen-val_man):.2f})")
                elif val_gen != val_man:
                    if (val_gen is None and val_man == 0) or (val_gen == 0 and val_man is None):
                        continue
                    mismatches += 1
                    print(f"  YTD Text Mismatch at Row {r} ({particulars_gen}) for {vert}: Generated={repr(val_gen)}, Manual={repr(val_man)}")
                    
        print(f"Summary for {cfg['name']}: {mismatches} mismatches found out of {total_checks} values compared.")
        if mismatches == 0:
            print(f"SUCCESS! All shared verticals for {cfg['name']} match perfectly!")
            
if __name__ == "__main__":
    compare_verticals()
