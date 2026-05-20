import openpyxl
import os

files = [
    "1. MIS_April 2025.xlsx",
    "MIS_May 2025.xlsx",
    "1. MIS_June 2025.xlsx"
]

for filename in files:
    path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", filename)
    # Load with formulas (data_only=False)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    # Load with values (data_only=True)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    
    sheet_f = wb_f['P&L']
    sheet_v = wb_v['P&L']
    
    print(f"=== {filename} ===")
    for r in range(1, 100):
        row_lbl = sheet_v.cell(row=r, column=1).value
        # If the row label looks like stock or purchase or COGS
        if row_lbl and any(k in str(row_lbl).lower() for k in ['stock', 'cogs', 'purchase', 'direct cost', 'material consumed']):
            print(f"Row {r} | Label: {row_lbl}")
            # Let's print values and formulas for columns L (Monthly Bluestreak), N (Monthly Clarus), X (YTD)
            # Wait, let's see which columns are used in April, May, June 2025.
            # Usually column L, N, P, etc. Let's print columns 10 to 30 for this row.
            cols_to_print = []
            for c in range(1, 30):
                val_v = sheet_v.cell(row=r, column=c).value
                val_f = sheet_f.cell(row=r, column=c).value
                if val_v is not None or val_f is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    cols_to_print.append(f"{col_letter}: [Val={val_v}, Formula={val_f}]")
            print("  " + "\n  ".join(cols_to_print[:10])) # limit output
    print()
