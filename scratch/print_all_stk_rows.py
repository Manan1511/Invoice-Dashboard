import openpyxl
import os

files = [
    "1. MIS_April 2025.xlsx",
    "MIS_May 2025.xlsx",
    "1. MIS_June 2025.xlsx"
]

for filename in files:
    path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"=== {filename} ===")
    if 'Stk ' in wb.sheetnames:
        sheet = wb['Stk ']
        print(f"Total rows: {sheet.max_row}")
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 8)]
            if any(x is not None for x in row_vals):
                # Search for 'Factory' or 'Closing Stock (Manual)' or 'closing stock'
                row_str = str(row_vals)
                if 'Factory' in row_str or 'Closing Stock' in row_str or 'Manual' in row_str or 'closing' in row_str.lower():
                    print(f"Row {r}: {row_vals}")
    print()
