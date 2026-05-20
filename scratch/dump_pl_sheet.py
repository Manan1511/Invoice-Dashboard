import openpyxl
import os

path = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/1. MIS_June 2025.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['P&L']

print("=== Manual Workbook P&L Lower Part ===")
for r in range(50, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    if any(x is not None for x in row_vals):
        # Format the row beautifully
        print(f"Row {r:<3} | {str(row_vals[0] or ''):<35} | {', '.join(str(v) if v is not None else '' for v in row_vals[1:8])}")
