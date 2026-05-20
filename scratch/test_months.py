import os
import sys
import shutil
import openpyxl

# Add backend directory to path
sys.path.append("C:/Users/manan/Downloads/Projects/Invoice Dashboard/backend")

from services.tb_parser import parse_tally_tb
from services.ytd_calculator import roll_forward_ytd
from services.workbook_builder import generate_monthly_workbook
from services.ledger_mapper import parse_ledger_file, replace_ledger_list_in_template

TEMPLATE_PATH = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/templates/MIS_template.xlsx"
BACKUP_PATH = "C:/Users/manan/Downloads/Projects/Invoice Dashboard/templates/MIS_template.xlsx.bak"

def force_excel_recalculate(file_path: str):
    """Launches headless Excel, opens the file, forces calculation, and saves it."""
    import win32com.client
    abs_path = os.path.abspath(file_path)
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(abs_path)
        excel.CalculateFull()
        wb.Save()
        wb.Close()
        print(f"  Headless Excel calculation completed for: {os.path.basename(file_path)}")
    except Exception as e:
        print(f"  [ERROR] headlessly recalculating workbook: {e}")
    finally:
        excel.Quit()

def compare_sheets(gen_path: str, gen_man_path: str, sheet_name: str, max_row: int = 100, max_col: int = 25) -> list:
    """Compares cells of a specific sheet between generated and manual workbooks."""
    wb_gen = openpyxl.load_workbook(gen_path, data_only=True)
    wb_man = openpyxl.load_workbook(gen_man_path, data_only=True)
    
    if sheet_name not in wb_gen.sheetnames or sheet_name not in wb_man.sheetnames:
        return [f"Sheet '{sheet_name}' not found in one of the workbooks."]
        
    sheet_gen = wb_gen[sheet_name]
    sheet_man = wb_man[sheet_name]
    
    mismatches = []
    
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val_gen = sheet_gen.cell(row=r, column=c).value
            val_man = sheet_man.cell(row=r, column=c).value
            
            # Match rules
            if isinstance(val_gen, (int, float)) and isinstance(val_man, (int, float)):
                if abs(val_gen - val_man) > 0.1:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    particulars = sheet_gen.cell(row=r, column=1).value or ""
                    mismatches.append(
                        f"Row {r}, Col {col_letter} ({particulars}): Generated={val_gen:.2f}, Manual={val_man:.2f} (Diff={abs(val_gen - val_man):.2f})"
                    )
            elif val_gen != val_man:
                # Treat None and 0/"" as equivalent to prevent noisy mismatches
                if val_gen is None and val_man is None:
                    continue
                if (val_gen is None and val_man == 0) or (val_gen == 0 and val_man is None):
                    continue
                if (val_gen is None and val_man == "") or (val_gen == "" and val_man is None):
                    continue
                    
                # For string, do case-insensitive whitespace strip match
                if isinstance(val_gen, str) and isinstance(val_man, str):
                    if val_gen.strip().lower() == val_gen.strip().lower(): # wait, val_gen and val_man comparison: let's fix it below
                        if val_gen.strip().lower() == val_man.strip().lower():
                            continue
                        
                col_letter = openpyxl.utils.get_column_letter(c)
                particulars = sheet_gen.cell(row=r, column=1).value or ""
                mismatches.append(
                    f"Row {r}, Col {col_letter} ({particulars}): Generated={repr(val_gen)}, Manual={repr(val_man)}"
                )
                
    return mismatches

def run_multi_month_test():
    # 1. Back up template
    if not os.path.exists(BACKUP_PATH):
        print(f"Creating template backup: {BACKUP_PATH}")
        shutil.copy(TEMPLATE_PATH, BACKUP_PATH)
        
    months_config = [
        {
            "name": "April 2025",
            "manual_file": "1. MIS_April 2025.xlsx",
            "output_file": "scratch/output_April_2025.xlsx",
            "month": 4,
            "year": 2025,
            "use_prior": False
        },
        {
            "name": "May 2025",
            "manual_file": "MIS_May 2025.xlsx",
            "output_file": "scratch/output_May_2025.xlsx",
            "month": 5,
            "year": 2025,
            "use_prior": True,
            "prior_file": "scratch/output_April_2025.xlsx"
        },
        {
            "name": "June 2025",
            "manual_file": "1. MIS_June 2025.xlsx",
            "output_file": "scratch/output_June_2025.xlsx",
            "month": 6,
            "year": 2025,
            "use_prior": True,
            "prior_file": "scratch/output_May_2025.xlsx"
        }
    ]
    
    results = {}
    
    try:
        for cfg in months_config:
            print(f"\n==========================================")
            print(f"--- Running Pipeline for {cfg['name']}")
            print(f"==========================================")
            
            manual_path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", cfg["manual_file"])
            output_path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", cfg["output_file"])
            prior_path = os.path.join("C:/Users/manan/Downloads/Projects/Invoice Dashboard", cfg["prior_file"]) if cfg["use_prior"] else None
            
            # Step A: Parse mappings from manual file and overwrite template mapping list
            print("Step A: Syncing mappings from manual sheet...")
            mappings = parse_ledger_file(manual_path)
            replace_ledger_list_in_template(mappings)
            print(f"  Successfully applied {len(mappings)} mappings to template.")
            
            # Step B: Parse Tally Trial Balance from manual sheet
            print("Step B: Parsing Trial Balance...")
            entries = parse_tally_tb(manual_path)
            print(f"  Parsed {len(entries)} ledger entries.")
            
            # Step C: Roll forward YTD
            print("Step C: Rolling forward YTD...")
            entries = roll_forward_ytd(entries, prior_path)
            
            # Step D: Generate monthly workbook
            print("Step D: Generating workbook...")
            generate_monthly_workbook(
                parsed_entries=entries,
                uploaded_file_path=manual_path,
                output_path=output_path,
                month=cfg["month"],
                year=cfg["year"],
                closing_stock=0.0
            )
            print(f"  Workbook generated: {cfg['output_file']}")
            
            # Step E: Force recalculate using headless Excel
            print("Step E: Forcing calculation via Headless Excel COM...")
            force_excel_recalculate(output_path)
            
            # Step F: Compare sheet values
            print("Step F: Comparing P&L and COGS sheets cell-by-cell...")
            pl_mismatches = compare_sheets(output_path, manual_path, "P&L", max_row=110, max_col=24)
            cogs_mismatches = compare_sheets(output_path, manual_path, "COGS", max_row=25, max_col=12)
            
            results[cfg["name"]] = {
                "pl_mismatches": pl_mismatches,
                "cogs_mismatches": cogs_mismatches
            }
            
            print(f"Results for {cfg['name']}:")
            print(f"  P&L mismatches: {len(pl_mismatches)}")
            if pl_mismatches:
                print("    Sample mismatches:")
                for m in pl_mismatches[:5]:
                    print(f"      {m}")
            print(f"  COGS mismatches: {len(cogs_mismatches)}")
            if cogs_mismatches:
                print("    Sample mismatches:")
                for m in cogs_mismatches[:5]:
                    print(f"      {m}")
                    
    finally:
        # Restore template
        if os.path.exists(BACKUP_PATH):
            print(f"\nRestoring template backup from {BACKUP_PATH}...")
            shutil.copy(BACKUP_PATH, TEMPLATE_PATH)
            os.remove(BACKUP_PATH)
            print("Template successfully restored.")
            
    print("\n==========================================")
    print("FINAL SUMMARY REPORT")
    print("==========================================")
    all_success = True
    for name, res in results.items():
        pl_len = len(res["pl_mismatches"])
        cogs_len = len(res["cogs_mismatches"])
        if pl_len == 0 and cogs_len == 0:
            print(f"[SUCCESS] {name}: 100% PERFECT MATCH! P&L and COGS math matches the manual CA report perfectly.")
        else:
            print(f"[FAIL] {name}: Failed with {pl_len} P&L mismatches and {cogs_len} COGS mismatches.")
            all_success = False
            
    if all_success:
        print("\nSUCCESS! All calculations match the manual CA reports perfectly across all three months!")
    else:
        print("\nWARNING: Calculations had mismatches. Please inspect sample mismatches above.")

if __name__ == "__main__":
    run_multi_month_test()
